import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import RED, BLACK
import re

# Existing Excel workbook
workbook = load_workbook("price.xlsx")
sheet = workbook.active

# Set column headers for the first two columns
sheet.cell(row=1, column=1).value = "Item Name"
sheet.cell(row=1, column=2).value = "Price yesterday"
sheet.cell(row=1, column=3).value = "Price"
sheet.cell(row=1, column=4).value = "Modified Price"  # New column header

# Set column header for "Price Yesterday" if it doesn't exist
if "Price Yesterday" not in [cell.value for cell in sheet[1]]:
    sheet.cell(row=1, column=2).value = "Price Yesterday"

# Get the maximum row index in the existing Excel file
max_row = sheet.max_row

# Create a font style for red color
font_red = Font(color=RED)
font_black = Font(color=BLACK)

row_idx = max_row + 1  # Starting row index for new data

urls = [
    "list of urls"
]

for url in urls:
    response = requests.get(url)
    html_content = response.text
    soup = BeautifulSoup(html_content, "html.parser")

    item_name_elements = soup.find_all(class_="item-name")
    price_elements = soup.find_all(class_="main-price")

    for item_name_element, price_element in zip(item_name_elements, price_elements):
        item_name = item_name_element.text.strip()
        price = price_element.text.strip()

        # Check if the item already exists in the Excel file
        found = False
        for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1):
            if row[0].value == item_name:
                found = True
                # Get the old price from "Price" column
                old_price = int(sheet.cell(row=row[0].row, column=3).value)
                # Set the "Price Yesterday" column to the old price
                sheet.cell(row=row[0].row, column=2).value = old_price
                new_price = int(re.sub(r'[^0-9.]', '', price))
                cell = sheet.cell(row=row[0].row, column=3)
                cell.value = (re.sub(r'[^0-9.]', '', price))
                if old_price != new_price:
                    cell.font = font_red
                else:
                    cell.font = font_black
                cell = sheet.cell(row=row[0].row, column=4)
                if new_price > 100000:
                    modified_price = new_price - 2000
                else:
                    modified_price = new_price - 1000
                cell.value = modified_price
                break

#        if not found:
            # Write new data to the Excel file
#            sheet.cell(row=row_idx, column=1).value = item_name
#            sheet.cell(row=row_idx, column=2).value = price
            #sheet.cell(row=row_idx, column=3).value = "New Entry"  # Indicate that it's a new entry
#            sheet.cell(row=row_idx, column=3).font = font_red
            #sheet.cell(row=row_idx, column=4).value = "New Entry"
#            row_idx += 1

# Save the modified Excel file
workbook.save("price.xlsx")
